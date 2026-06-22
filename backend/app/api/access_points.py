from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from typing import List, Optional
from app.database import get_db
from app.models import AccessPoint, Project
from app.schemas import (
    AccessPointCreate, AccessPointUpdate, AccessPointMove,
    AccessPointResponse, ImportResult,
)
from app.models import Stage
from pydantic import BaseModel
import openpyxl
import re
import io


class TextParseRequest(BaseModel):
    text: str

router = APIRouter()


@router.get("/projects/{project_id}/access-points", response_model=List[AccessPointResponse])
def list_access_points(
    project_id: int,
    stage_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(AccessPoint).filter(AccessPoint.project_id == project_id)
    if stage_id is not None:
        q = q.filter(AccessPoint.stage_id == stage_id)
    return q.order_by(AccessPoint.sort_order, AccessPoint.created_at).all()


@router.post("/access-points", response_model=AccessPointResponse)
def create_access_point(data: AccessPointCreate, db: Session = Depends(get_db)):
    ap = AccessPoint(**data.dict())
    db.add(ap)
    db.commit()
    db.refresh(ap)
    return ap


@router.put("/access-points/{ap_id}", response_model=AccessPointResponse)
def update_access_point(ap_id: int, data: AccessPointUpdate, db: Session = Depends(get_db)):
    ap = db.query(AccessPoint).filter(AccessPoint.id == ap_id).first()
    if not ap:
        raise HTTPException(status_code=404, detail="接入点不存在")
    update_data = data.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(ap, key, value)
    db.commit()
    db.refresh(ap)
    return ap


@router.delete("/access-points/{ap_id}")
def delete_access_point(ap_id: int, db: Session = Depends(get_db)):
    ap = db.query(AccessPoint).filter(AccessPoint.id == ap_id).first()
    if not ap:
        raise HTTPException(status_code=404, detail="接入点不存在")
    db.delete(ap)
    db.commit()
    return {"message": "已删除"}


@router.put("/access-points/{ap_id}/move", response_model=AccessPointResponse)
def move_access_point(ap_id: int, data: AccessPointMove, db: Session = Depends(get_db)):
    """拖拽移动：改 stage_id + sort_order"""
    ap = db.query(AccessPoint).filter(AccessPoint.id == ap_id).first()
    if not ap:
        raise HTTPException(status_code=404, detail="接入点不存在")
    ap.stage_id = data.stage_id
    ap.sort_order = data.sort_order
    db.commit()
    db.refresh(ap)
    return ap


@router.post("/projects/{project_id}/access-points/upload", response_model=ImportResult)
def upload_access_points(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # 解析表头映射
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value or "").strip()
        if val in ("名称", "接入点名称", "网点名称"):
            headers["name"] = col
        elif val in ("地址", "安装地址", "详细地址"):
            headers["address"] = col

    if "name" not in headers:
        return ImportResult(imported=0, skipped=0, skipped_details=["未找到名称列（名称/接入点名称/网点名称）"])

    # 创建接入点
    imported = 0
    skipped = 0
    skipped_details = []
    existing_names = {
        ap.name for ap in db.query(AccessPoint).filter(AccessPoint.project_id == project_id).all()
    }
    seen_names = set()

    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row, column=headers["name"]).value or "").strip()
        if not name:
            continue

        if name in seen_names:
            skipped += 1
            skipped_details.append(f"第{row}行：{name}（批次内重复）")
            continue

        if name in existing_names:
            skipped += 1
            skipped_details.append(f"第{row}行：{name}（数据库中已存在）")
            continue

        address = ""
        if "address" in headers:
            address = str(ws.cell(row=row, column=headers["address"]).value or "").strip()

        ap = AccessPoint(project_id=project_id, name=name, address=address)
        db.add(ap)
        seen_names.add(name)
        imported += 1

    db.commit()
    return ImportResult(imported=imported, skipped=skipped, skipped_details=skipped_details[:20])


@router.post("/projects/{project_id}/access-points/parse-text", response_model=ImportResult)
def parse_text_report(project_id: int, data: TextParseRequest, db: Session = Depends(get_db)):
    """解析自由文本进度汇报，自动创建接入点并分配到对应阶段"""
    text = data.text
    imported = 0
    skipped = 0
    skipped_details = []

    existing_names = {
        ap.name for ap in db.query(AccessPoint).filter(AccessPoint.project_id == project_id).all()
    }
    stages = {s.name: s for s in db.query(Stage).filter(Stage.project_id == project_id).all()}

    # 按 -- 拆分各阶段
    parts = re.split(r'--|——', text)

    for part in parts:
        # 提取阶段名和接入点列表
        m = re.search(r'[①②③④⑤⑥⑦⑧⑨⑩](\S+?)（\d+个）[：:](.+)', part)
        if not m:
            continue

        stage_name = m.group(1).strip()
        ap_text = m.group(2).strip().rstrip('。')

        # 查找或创建阶段
        stage = stages.get(stage_name)
        if not stage:
            # 尝试模糊匹配
            for sname, s in stages.items():
                if stage_name in sname or sname in stage_name:
                    stage = s
                    break
        if not stage:
            # 创建新阶段
            max_order = max((s.sort_order for s in stages.values()), default=-1)
            is_bottleneck = 1 if '瓶颈' in stage_name else 0
            color = '#ff4d4f' if is_bottleneck else '#1890ff'
            stage = Stage(
                project_id=project_id, name=stage_name,
                sort_order=max_order + 1, is_bottleneck=is_bottleneck, color=color,
            )
            db.add(stage)
            db.flush()
            stages[stage_name] = stage

        # 解析接入点名称
        # 用 、 或 ）结尾来拆分
        # 格式：名称1（备注）、名称2、名称3（备注）。
        ap_names = []
        current = ''
        depth = 0
        for ch in ap_text:
            if ch == '（' or ch == '(':
                depth += 1
            elif ch == '）' or ch == ')':
                depth -= 1
            elif ch in ('、', '。', '，') and depth == 0:
                name = current.strip()
                # 提取纯名称（去掉后面的括号备注作为地址）
                addr = ''
                m2 = re.match(r'^(.+?)（(.+?)）$', name)
                if m2:
                    name = m2.group(1).strip()
                    addr = m2.group(2).strip()
                if name and name not in ('个', ''):
                    ap_names.append((name, addr))
                current = ''
                continue
            current += ch

        # 最后一个
        name = current.strip()
        addr = ''
        m2 = re.match(r'^(.+?)（(.+?)）$', name)
        if m2:
            name = m2.group(1).strip()
            addr = m2.group(2).strip()
        if name and name not in ('个', ''):
            ap_names.append((name, addr))

        # 批量创建
        for ap_name, ap_addr in ap_names:
            if ap_name in existing_names:
                skipped += 1
                skipped_details.append(f"{ap_name}（已存在）")
                continue
            ap = AccessPoint(
                project_id=project_id, stage_id=stage.id,
                name=ap_name, address=ap_addr,
            )
            db.add(ap)
            existing_names.add(ap_name)
            imported += 1

    db.commit()
    return ImportResult(imported=imported, skipped=skipped, skipped_details=skipped_details[:20])
